import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook

import GridCal.Engine as gc

from GridCal.Engine.IO.file_handler import FileOpen
from GridCal.Engine.Simulations.PowerFlow.power_flow_worker import PowerFlowOptions
from GridCal.Engine.Simulations.PowerFlow.power_flow_options import ReactivePowerControlMode, SolverType
from GridCal.Engine.Simulations.PowerFlow.power_flow_driver import PowerFlowDriver

from sklearn.metrics import r2_score


#%%

main_circuit = FileOpen('IEEE118busREE_Winter.raw').open()

# for bus in main_circuit.buses:
#     if bus.determine_bus_type()._value_==1:
#         bus.Vm0=1
#         bus.Va0=0
#     elif bus.determine_bus_type()._value_==2:
#         bus.controlled_generators[0].Vset=1
    

#%%

generators=main_circuit.get_generators()

#%%
for gen in generators:
    bus_name=gen.bus._name
    #Snom=Generators.loc[Generators.query('BusName == @bus_name').index[0],'TOT']
    gen.Snom=gen.Snom*3
    gen.P=gen.P*2
    # gen.Pmax=Generators.loc[Generators.query('BusName == @bus_name').index[0],'TOT']
    # gen.Pmin=0
    gen.Qmax=gen.Qmax*3
    #0.9*Generators.loc[Generators.query('BusName == @bus_name').index[0],'TOT']
    gen.Qmin=gen.Qmin*3
    #gen.is_controlled=True
    
#%%

loads= main_circuit.get_loads()
buses=main_circuit.get_buses()
buses[44].loads=[]
for load in loads:
    load.B=load.B*2
    load.Ir=load.Ir*2
    


#%%
# Lines=pd.read_csv('G:/Il mio Drive/Francesca 118 v2/additional-files-mti-118/additional-files-mti-118/Lines.csv')
# for i in range(0,len(Lines)):
#     Lines.loc[i,'FB_Num']=int(Lines.loc[i,'Bus from '][3:])
#     Lines.loc[i,'TB_Num']=int(Lines.loc[i,'Bus to'][3:])

# for branch in main_circuit.get_branches():    
#     if branch.type_name=='Line':
#         from_bus=int(branch.bus_from.code)
#         to_bus=int(branch.bus_to.code)
        
#         branch.X=Lines.loc[Lines.query('FB_Num == @from_bus and TB_Num == @to_bus').index[0],'Reactance (p.u.)']
#         branch.R=Lines.loc[Lines.query('FB_Num == @from_bus and TB_Num == @to_bus').index[0],'Resistance (p.u.)']
#         #branch.B=0

# #%%        
# for branch in main_circuit.get_branches():    
#     if branch.type_name=='Transformer':
#         #print(branch.X)
#         branch.X=0.01

#%%

#for solver_type in [SolverType.NR]:#, SolverType.IWAMOTO, SolverType.LM, SolverType.FASTDECOUPLED]:

solver_type=SolverType.NR
    
options = PowerFlowOptions(solver_type,
                           verbose=False,
                           initialize_with_existing_solution=False,
                           multi_core=False,
                           dispatch_storage=True,
                           control_q=ReactivePowerControlMode.Direct,
                           control_p=False,
                           retry_with_other_methods=False)


power_flow = PowerFlowDriver(main_circuit, options)
    # power_flow.run()
    # res=power_flow.results
    
    # print(res.convergence_reports[0].converged_)


#%%

# for bus in power_flow.grid.buses:
#     if bus.determine_bus_type()._value_!=3:
#         bus.determine_bus_type()._value_=1

# i=0
# buses=power_flow.grid.buses
# for bus in power_flow.grid.buses:
#     if bus.determine_bus_type()._value_==2:
#         print(i)
#         print(bus.name)
#     i=i+1
#         #bus.determine_bus_type()._value_=2


#%%
power_flow.run()

res=power_flow.results

v = np.abs(power_flow.results.voltage)
va = np.angle(power_flow.results.voltage,deg=True)

print(res.convergence_reports[0].converged_)

#%%

def process_GridCal_PF(GridCal_grid, pf_results):
    
    Sbase = GridCal_grid.Sbase
    
    # pf_bus: voltage and angle in buses
    
    bus = [int(bus) for bus in [bus.code for bus in GridCal_grid.buses]]
    Vm = np.abs(pf_results.results.voltage)
    theta = np.angle(pf_results.results.voltage, deg=True) 
    pf_bus = pd.DataFrame({'bus':bus, 'Vm': Vm, 'theta':theta, 'type':pf_results.results.bus_types})
            
    # pf_load: active and reactive power in loads
    
    bus_load = [int(load.bus.code) for load in GridCal_grid.get_loads()]   
    idx_load = [i for i, bus in enumerate(bus) if bus in bus_load]   
    Vm = np.abs([pf_results.results.voltage[idx] for idx in idx_load])
    theta = np.angle([pf_results.results.voltage[idx] for idx in idx_load], deg=True)     
    P = np.array([pf_results.grid.buses[idx].loads[0].Ir for idx in idx_load])/Sbase
    Q = -np.array([pf_results.grid.buses[idx].loads[0].B for idx in idx_load])/Sbase
    
    # B = np.array([pf_results.grid.buses[idx].loads[0].B for idx in idx_load])
    # Ir = np.array([pf_results.grid.buses[idx].loads[0].Ir for idx in idx_load])
    # P = Ir
    # Q=abs(B)
    
    pf_load = pd.DataFrame({'bus':bus_load, 'Vm':Vm, 'theta':theta, 'P':P, 'Q':Q})
    
    #pf_gen: active and reactive power in generator buses
    
    bus_gen = [int(gen.bus.code) for gen in GridCal_grid.get_generators()]   
    idx_gen = [i for i, bus in enumerate(bus) if bus in bus_gen]   
    Vm = np.abs([pf_results.results.voltage[idx] for idx in idx_gen])
    theta = np.angle([pf_results.results.voltage[idx] for idx in idx_gen], deg=True) 
    P = [(np.real(pf_results.results.Sbus[idx]) + (pf_results.grid.buses[idx].loads[0].Ir if bus[idx] in bus_load else 0))/Sbase for idx in idx_gen]
    Q = [(np.imag(pf_results.results.Sbus[idx]) - (pf_results.grid.buses[idx].loads[0].B if bus[idx] in bus_load else 0))/Sbase for idx in idx_gen]

    Qmin= [gen.Qmin/Sbase for gen in GridCal_grid.get_generators()]  
    Qmax= [gen.Qmax/Sbase for gen in GridCal_grid.get_generators()]  
    Pmin= [gen.Pmin/Sbase for gen in GridCal_grid.get_generators()]  
    Pmax= [gen.Pmax/Sbase for gen in GridCal_grid.get_generators()]  
    pf_gen = pd.DataFrame({'bus':bus_gen, 'Vm':Vm, 'theta':theta, 'P':P, 'Q':Q,'Qmin':Qmin,'Qmax':Qmax,'Pmin':Pmin,'Pmax':Pmax})
    
    return pf_bus, pf_load, pf_gen

pf_bus, pf_load, pf_gen=process_GridCal_PF(main_circuit, power_flow)

#%%
pf_load=pf_load.query('P!=0')

#%%

def overwrite_sheet(book,sheet_name):
    # Check if the sheet exists in the workbook
    #if sheet_name in book.sheetnames:
    sheet = book[sheet_name]

    # Clear existing content in the sheet
    sheet.delete_rows(1, sheet.max_row)  # This clears all rows, adjust as needed


    # Convert the DataFrame to an openpyxl worksheet
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r_idx, row in enumerate(dataframe_to_rows(df_existing, index=False,header=1), sheet.max_row + 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)        
    return book

#%%

# Load the existing Excel file into a DataFrame
existing_file = 'G:/Il mio Drive/Francesca 118 v2/ree_orig/tool_ree/tool_ree/01_data/IEEE_118bus_full_GFOLSLACK_6a.xlsx'
df_existing = pd.read_excel(existing_file,header=1,sheet_name='PF')

df_existing['Vm']=v
df_existing['theta']=va

# Specify the name of the sheet you want to overwrite
sheet_name = 'PF'

# Load the existing Excel file using openpyxl
book = load_workbook(existing_file)

book= overwrite_sheet(book, sheet_name)
# Save the changes back to the Excel file
book.save(existing_file)
    
#%%

def calculate_R_X_loads(T_load, pf_load):
    
    T_load['V'] = pf_load["Vm"]
    T_load['theta'] = pf_load["theta"]
    T_load['P'] = pf_load['P']
    T_load['Q'] = pf_load["Q"]
    # T_load.loc[T_load["type"] == "PQ", "R"] = T_load.loc[T_load["type"] == "PQ", "V"] ** 2 / T_load.loc[T_load["type"] == "PQ", "P"]
    # T_load.loc[T_load["type"] == "PQ", "X"] = T_load.loc[T_load["type"] == "PQ", "V"] ** 2 / T_load.loc[T_load["type"] == "PQ", "Q"]  
    T_load["R"] = T_load["V"] ** 2 / T_load["P"]
    T_load["X"] = T_load["V"] ** 2 / T_load["Q"]
    return T_load

#%%
sheet_name = 'load'
df_existing = pd.DataFrame()#pd.read_excel(existing_file,sheet_name=sheet_name,header=1)
#df_existing.columns=['number', 'bus', 'R', 'X', 'P', 'Q']
df_existing['bus']=pf_load['bus']


df_existing=calculate_R_X_loads(df_existing,pf_load)
df_existing['number']=np.arange(1,len(df_existing)+1)

df_existing=df_existing.reindex(['number', 'bus', 'R', 'X', 'P', 'Q'],axis=1)

# Load the existing Excel file using openpyxl
book = load_workbook(existing_file)


book= overwrite_sheet(book, sheet_name)
# Save the changes back to the Excel file
book.save(existing_file)

#%% 
sheet_name='CASE'
df_existing = pd.read_excel(existing_file,sheet_name=sheet_name,header=1)

df_existing['P']=pf_gen['P']
df_existing['Q']=pf_gen['Q']
df_existing['V']=pf_gen['Vm']
# df_existing['S']=Generators['Snom']/main_circuit.Sbase

# df_existing['S_SG']=Generators['S_SG']/main_circuit.Sbase
# df_existing['S_GFOR']=Generators['S_GFOR']/main_circuit.Sbase
# df_existing['S_GFOL']=Generators['S_GFOL']/main_circuit.Sbase

# pf=0.8
# tg_phi=np.tan(np.arccos(pf))

# df_existing['P_SG']=Generators['P_SG']/main_circuit.Sbase
# df_existing['P_GFOR']=Generators['P_GFOR']/main_circuit.Sbase
# df_existing['P_GFOL']=Generators['P_GFOL']/main_circuit.Sbase

# df_existing['Q_SG']=tg_phi*Generators['P_SG']/main_circuit.Sbase
# df_existing['Q_GFOR']=tg_phi*Generators['P_GFOR']/main_circuit.Sbase
# df_existing['Q_GFOL']=tg_phi*Generators['P_GFOL']/main_circuit.Sbase

# Load the existing Excel file using openpyxl
book = load_workbook(existing_file)


book= overwrite_sheet(book, sheet_name)
# Save the changes back to the Excel file
book.save(existing_file)

#%%

#df_existing.loc[0,['number', 'bus', 'P', 'Q', 'V', 'delta', 'type', 'element', 'Sb', 'Vb']]=1

sheet_name='user'
df_existing = pd.read_excel(existing_file,sheet_name=sheet_name,header=1)

for gen in generators:
    bus_num=int(gen.bus.code)
    df_existing.loc[df_existing.query('bus == @bus_num').index,'Sb']=gen.Snom

book = load_workbook(existing_file)


book= overwrite_sheet(book, sheet_name)
# Save the changes back to the Excel file
book.save(existing_file)




# #%%
# main_circuit2 = FileOpen('IEEE118busREE_Winter.raw').open()
# generators=main_circuit2.get_generators()

# for gen in generators:
    
#     if gen.bus.determine_bus_type()._value_ != 3:
#         gen.bus.determine_bus_type()._value_ =1
#         #gen.is_controlled=False
        
#     # bus=gen.bus    
    
#     # if bus.determine_bus_type()._value_ == 3:
#     #     continue
#     # else:
#     #     bus_code=int(gen.bus.code)
    
#     #     P=float(pf_gen.loc[pf_gen.query('bus == @bus_code').index,'P'])*100
#     #     Q=float(pf_gen.loc[pf_gen.query('bus == @bus_code').index,'Q'])*100
        
#     #     bus.controlled_generators=[]
#     #     main_circuit2.add_load(bus, gc.Load('gen', P=-P, Q=-Q))

# #%%

# power_flow2 = PowerFlowDriver(main_circuit2, options)

# power_flow2.run()

# res2=power_flow2.results

# v = np.abs(power_flow2.results.voltage)
# va = np.angle(power_flow2.results.voltage,deg=True)

# print(res2.convergence_reports[0].converged_)

# #%%
# generators=main_circuit.get_generators()
# for gen in generators:
    


# #%%



